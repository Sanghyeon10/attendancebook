import time
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import datetime
import making
import re
import matplotlib.pyplot as plt
from openpyxl import Workbook
import making
import os


filename="2025년 초등2부 목장 편성표"
grouplist=['4-1', '4-2', '4-3', '4-4', '5-1', '5-2', '5-3', '5-4', '5-5', '6-1', '6-2', '6-3', '6-4','6-5']



# 데이터를 불러오는 함수 (예제 텍스트 파일 이름)
def load_data():
    # 목장과 선생님 이름
    ranch_teacher= making.get_name()

    # 선생님 이름과 전화번호
    teacher_phone = making.get_phonenumber()

    # file_list = os.listdir(os.getcwd())
    # print(os.getcwd())
    # print(file_list)
    # 목장과 그 목장 아이들 리스트
    ranch_students =making.get_keyAndlist('farmnameAndkids')


    # 아이 이름과 출결
    student_information = pd.read_excel(r"C:\Users\captu\Downloads\아이들 정보.xlsx", sheet_name='시트1' , index_col=0)

    return ranch_teacher, teacher_phone, ranch_students, student_information





# 엑셀 파일 생성 함수
def create_excel():
    ranch_teacher, teacher_phone, ranch_students, student_information = load_data()
    print(load_data())
    workbook = Workbook()

    for ranch, teacher in ranch_teacher.items():
        # 새로운 시트 생성
        sheet = workbook.create_sheet(title=ranch)

        # 선생님 정보 작성
        teacher_info = f"{ranch} 목장: {teacher}선생님 ({teacher_phone[teacher]})"
        sheet.cell(row=2, column=1, value=teacher_info) #그래야지만, 1,1은 숨기기해서 안보여짐.

        # 헤더 작성
        headers = ["번호", "이름", "학생연락처", "부모님 연락처", "출결", "생일", "비고"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=3, column=col_num, value=header)

        # 아이들 정보 작성
        students = ranch_students[ranch]
        for idx, student in enumerate(students, 1):
            sheet.cell(row=idx + 3, column=1, value=idx)  # 번호
            sheet.cell(row=idx + 3, column=2, value=student)  # 이름
            try:
                sheet.cell(row=idx + 3, column=3, value=student_information.loc[student,student_information.columns[0]])  # 전화번호
                sheet.cell(row=idx + 3, column=4, value=student_information.loc[student,student_information.columns[1]])  # 부모님 전화번호
                sheet.cell(row=idx + 3, column=5, value=student_information.loc[student,student_information.columns[2]])  #출결
                sheet.cell(row=idx + 3, column=6, value=student_information.loc[student, student_information.columns[3]]) # 생일
                sheet.cell(row=idx + 3, column=7, value=student_information.loc[student, student_information.columns[4]]) #비고

                # time.sleep(1)
            except:
                pass
            print(ranch, student)


    # 기본 시트 제거
    default_sheet = workbook["Sheet"]
    workbook.remove(default_sheet)

    # 엑셀 파일 저장
    workbook.save("{}.xlsx".format(filename))

def save_each_sheet_as_file(input_file):
    try:
        # 모든 시트 읽기
        sheets = pd.read_excel(input_file, sheet_name=None)  # 모든 시트를 읽어옴

        # 각 시트를 별도 파일로 저장
        for sheet_name, df in sheets.items():
            # 출력 파일 경로 생성

            df.to_excel("{}.xlsx".format(sheet_name))  # 5-1식으로 출력
            # 출력해야지 그 출력된 결과물에서 다시 통계작성함.
            print(f"시트 '{sheet_name}'이(가) 파일로 저장되었습니다: ")
    except Exception as e:
        print(f"오류 발생: {e}")


# 실행
create_excel()
save_each_sheet_as_file(filename+".xlsx")


scopes = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/dirve'
]
creds = ServiceAccountCredentials.from_json_keyfile_name(making.addrresOfjsonfile)
#위치 바뀌면 수정해줄것.

file = gspread.authorize(creds)
# sh = file.open("오류 테스트용")
sh = file.open(filename) #woorbook = sh
# print(type(sh.worksheets()[1]))

A = input('특정 목장 업로드할지 판단 0이면 전부올리기 input')
uploadlist = making.WhatIsToUpload(A)
print("uploadlist:",uploadlist)


all_group= grouplist
worksheet_list = sh.worksheets()

if len(worksheet_list)!= len(all_group): #에러상황일수도? 워크 시트 기준으로해야맞음
    print(worksheet_list)
    print(all_group)
    input('시트와 변수개수가 안맞다!')
    # time.sleep(600)

for i in range(len(all_group)):
    if all_group[i] in uploadlist: #업로드 리스트에 해당하는 것만 업로드하기
        try:
            sheet= sh.worksheet(all_group[i]) #구글 스프레드기준 찾기
            print(all_group[i])

            tempdf = pd.read_excel(r'{}.xlsx'.format(all_group[i])) #해당파일찾고 데이터 옮겨오기
            tempdf = tempdf.fillna('')

            sheet.clear()
            sheet.update(range_name=making.getrangename(tempdf) ,values=[tempdf.columns.values.tolist()] +tempdf.values.tolist()) #데이터 덧씌우기
            #6.0.0 버전되면 구문 위치 바뀐다고함.
            time.sleep(5)

            # print(making.getrangename(tempdf))
        except:
            print( all_group[i], "에러")
    else:
        pass

